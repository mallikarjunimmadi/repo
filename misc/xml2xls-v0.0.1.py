#!/usr/bin/env python3
import argparse
from pathlib import Path
import re
from collections import defaultdict
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


# ---------------- XML flattening ----------------
def strip_ns(tag: str) -> str:
    if tag.startswith("{") and "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def flatten_xml(elem, prefix, out, siblings_counter):
    """
    Flattens XML to key/value pairs:
      - element text -> path
      - attributes -> path/@attr
      - repeated siblings indexed like tag[1], tag[2]...
    """
    tag = strip_ns(elem.tag)
    base_path = f"{prefix}/{tag}" if prefix else tag

    key_for_count = (prefix, tag)
    idx = siblings_counter[key_for_count]
    siblings_counter[key_for_count] += 1
    path = f"{base_path}[{idx}]" if idx > 0 else base_path

    # attributes
    for k, v in elem.attrib.items():
        out[f"{path}/@{strip_ns(k)}"] = v

    # text
    text = (elem.text or "").strip()
    if text:
        out[path] = text

    # children
    child_counter = defaultdict(int)
    for child in list(elem):
        flatten_xml(child, path, out, child_counter)


def parse_file_as_xml_kv(path: Path) -> dict:
    tree = ET.parse(path)
    root = tree.getroot()
    out = {}
    siblings_counter = defaultdict(int)
    flatten_xml(root, "", out, siblings_counter)
    return out


# ---------------- Excel helpers ----------------
def sanitize_sheet_name(name: str) -> str:
    # max 31 chars, no: : \ / ? * [ ]
    name = re.sub(r'[:\\/?*\[\]]', "_", name).strip()
    return (name or "Sheet")[:31]


def safe_table_name(name: str) -> str:
    # Excel table name: letters, numbers, underscore; must start with letter/underscore
    name = re.sub(r"\W+", "_", name)
    if not name:
        name = "Tbl"
    if not re.match(r"^[A-Za-z_]", name):
        name = "_" + name
    return name[:60]


def autosize(ws, cols=None, max_width=120):
    if cols is None:
        cols = range(1, ws.max_column + 1)
    for col in cols:
        max_len = 0
        for cell in ws[get_column_letter(col)]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), max_width)


def add_table(ws, display_name: str):
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tab = Table(displayName=safe_table_name(display_name), ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def add_kv_sheet(wb, title, kv_items):
    """
    Per-file sheet: columns key,value
    """
    ws = wb.create_sheet(title=title)
    ws["A1"] = "key"
    ws["B1"] = "value"
    ws["A1"].font = ws["B1"].font = Font(bold=True)
    ws["A1"].alignment = ws["B1"].alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    r = 2
    for k, v in kv_items:
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
        r += 1

    autosize(ws, cols=[1, 2], max_width=120)
    if ws.max_row > 1:
        add_table(ws, f"{title}_tbl")
    return ws


def add_rows_sheet(wb, title, headers, rows, index=None, max_width=120):
    """
    Generic sheet with header row + many rows.
    """
    ws = wb.create_sheet(title=title, index=index)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append(list(row))

    autosize(ws, max_width=max_width)
    if ws.max_row > 1:
        add_table(ws, f"{title}_tbl")
    return ws


# ---------------- Main ----------------
def main():
    ap = argparse.ArgumentParser(
        description="One Excel output: per-file sheets + consolidated + per-extension sheets + common/not-common comparison."
    )
    ap.add_argument("--root", required=True, help="Directory containing .vpf/.xml files")
    ap.add_argument("--ext", action="append", default=["vpf", "xml"],
                    help="Extensions to include (repeatable). Default: vpf, xml")
    ap.add_argument("--recursive", action="store_true", help="Recurse into subdirectories")
    ap.add_argument("--out", default="report.xlsx", help="Output xlsx path")
    ap.add_argument("--skip-nonxml", action="store_true",
                    help="Skip files that fail XML parse (otherwise listed in Errors sheet)")
    args = ap.parse_args()

    root = Path(args.root).resolve()
    out_path = Path(args.out).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    exts = [e.lower().lstrip(".") for e in args.ext]

    if not root.exists():
        raise SystemExit(f"[ERROR] root not found: {root}")

    pattern = "**/*" if args.recursive else "*"
    files = sorted(
        [p for p in root.glob(pattern)
         if p.is_file() and p.suffix.lower().lstrip(".") in exts]
    )

    # Parse all files -> file_kv
    file_kv = {}      # rel -> dict(key->value)
    file_ext = {}     # rel -> ext
    errors = []       # (rel, error)

    for p in files:
        rel = str(p.relative_to(root))
        ext = p.suffix.lower().lstrip(".") or "noext"
        try:
            kv = parse_file_as_xml_kv(p)
            if kv:
                file_kv[rel] = kv
                file_ext[rel] = ext
            else:
                errors.append((rel, "Parsed but produced 0 key/value pairs"))
        except Exception as e:
            if args.skip_nonxml:
                continue
            errors.append((rel, str(e)))

    files_used = sorted(file_kv.keys())
    n_files_ok = len(files_used)

    # Build consolidated rows and per-extension rows
    consolidated_rows = []              # (file, ext, key, value)
    rows_by_ext = defaultdict(list)     # ext -> list of rows
    kv_rows_count = 0

    for f in files_used:
        ext = file_ext.get(f, "")
        kv_items = sorted(file_kv[f].items(), key=lambda x: x[0])
        for k, v in kv_items:
            consolidated_rows.append((f, ext, k, v))
            rows_by_ext[ext].append((f, k, v))
            kv_rows_count += 1

    # STRICT Common / Not_Common across all parsed files
    common_items = []      # (key, value)
    not_common_rows = []   # (key, file, value, status)

    if n_files_ok > 0:
        all_keys = set()
        for f in files_used:
            all_keys.update(file_kv[f].keys())

        for k in sorted(all_keys):
            values = []
            missing = False

            for f in files_used:
                v = file_kv[f].get(k)
                if v is None:
                    missing = True
                values.append(v)

            if missing:
                for f in files_used:
                    v = file_kv[f].get(k)
                    if v is None:
                        not_common_rows.append((k, f, "", "MISSING"))
                    else:
                        # not identical across ALL, so treat present as DIFFERENT for strict definition
                        not_common_rows.append((k, f, v, "DIFFERENT"))
                continue

            first = values[0]
            if all(v == first for v in values):
                common_items.append((k, first))
            else:
                for f in files_used:
                    not_common_rows.append((k, f, file_kv[f].get(k, ""), "DIFFERENT"))

    # ---------------- Write workbook ----------------
    wb = Workbook()
    wb.remove(wb.active)

    # Summary first
    summary_rows = [(
        len(files),
        n_files_ok,
        len(errors),
        kv_rows_count,
        len(common_items),
        len(not_common_rows),
        ", ".join(sorted(set(file_ext.values()))) if file_ext else ""
    )]
    add_rows_sheet(
        wb,
        title="Summary",
        headers=["files_found", "files_parsed_ok", "files_failed_or_empty", "total_kv_rows",
                 "common_pairs", "not_common_rows", "extensions_seen"],
        rows=summary_rows,
        index=0,
        max_width=60
    )

    # Consolidated
    add_rows_sheet(
        wb,
        title="Consolidated",
        headers=["file", "ext", "key", "value"],
        rows=consolidated_rows,
        index=1,
        max_width=120
    )

    # Common
    add_rows_sheet(
        wb,
        title="Common",
        headers=["key", "value"],
        rows=common_items,
        index=2,
        max_width=120
    )

    # Not_Common (sorted for readability)
    not_common_rows.sort(key=lambda x: (x[0], x[1]))
    add_rows_sheet(
        wb,
        title="Not_Common",
        headers=["key", "file", "value", "status"],
        rows=not_common_rows,
        index=3,
        max_width=120
    )

    # Per-extension sheets: EXT_vpf, EXT_xml, ...
    # Keep them near the top but after the main sheets
    ext_titles = []
    for ext in sorted(rows_by_ext.keys()):
        title = sanitize_sheet_name(f"EXT_{ext}")
        ext_titles.append(title)
        add_rows_sheet(
            wb,
            title=title,
            headers=["file", "key", "value"],
            rows=rows_by_ext[ext],
            index=len(wb.worksheets),  # append at end
            max_width=120
        )

    # Per-file sheets (may be many; but you have 5 now)
    # Note: Excel has a practical sheet limit; this is fine for small/medium sets.
    used_names = set(ws.title for ws in wb.worksheets)

    for f in files_used:
        stem = Path(f).stem  # from relative file path
        base = sanitize_sheet_name(stem)
        name = base
        i = 1
        while name in used_names:
            suffix = f"_{i}"
            name = sanitize_sheet_name(base[:31 - len(suffix)] + suffix)
            i += 1
        used_names.add(name)

        kv_items = sorted(file_kv[f].items(), key=lambda x: x[0])
        add_kv_sheet(wb, name, kv_items)

    # Errors
    if errors:
        add_rows_sheet(
            wb,
            title="Errors",
            headers=["file", "error"],
            rows=errors,
            index=len(wb.worksheets),  # append at end
            max_width=140
        )

    wb.save(out_path)
    print(f"[OK] Wrote {out_path} (files={n_files_ok}, kv_rows={kv_rows_count}, errors={len(errors)})")


if __name__ == "__main__":
    main()